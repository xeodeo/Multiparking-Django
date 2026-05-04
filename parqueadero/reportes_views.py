"""
Vistas para el módulo de Reportes y Estadísticas
"""
from django.shortcuts import render
from django.views import View
from django.http import HttpResponse
from django.db.models import Sum
from django.utils import timezone
from datetime import datetime, timedelta
from decimal import Decimal
import json

from usuarios.mixins import AdminRequiredMixin
from parqueadero.models import Piso, Espacio, InventarioParqueo
from pagos.models import Pago
from reservas.models import Reserva


class ReportesView(AdminRequiredMixin, View):
    """
    Vista principal de Reportes y Estadísticas con gráficas interactivas
    """

    def get(self, request):
        # Filtros
        periodo = request.GET.get('periodo', 'mes')  # mes, semana, año
        piso_id = request.GET.get('piso', 'todos')

        # Fechas según período
        now = timezone.now()
        if periodo == 'semana':
            fecha_inicio = now - timedelta(days=7)
        elif periodo == 'mes':
            fecha_inicio = now - timedelta(days=30)
        elif periodo == 'año':
            fecha_inicio = now - timedelta(days=365)
        else:
            fecha_inicio = now - timedelta(days=30)

        # Query base de inventario parqueo
        inventario_qs = InventarioParqueo.objects.filter(
            parHoraEntrada__gte=fecha_inicio
        )

        # Filtro por piso
        if piso_id != 'todos':
            inventario_qs = inventario_qs.filter(
                fkIdEspacio__fkIdPiso__pk=int(piso_id)
            )

        # ══════════════════════════════════════════
        # 1. MÉTRICAS PRINCIPALES
        # ══════════════════════════════════════════

        # Total espacios
        total_espacios = Espacio.objects.filter(espEstado__in=['DISPONIBLE', 'OCUPADO']).count()

        # Espacios actualmente ocupados
        espacios_ocupados = Espacio.objects.filter(espEstado='OCUPADO').count()

        # Ocupación promedio en el período
        if total_espacios > 0:
            ocupacion_promedio = int((espacios_ocupados / total_espacios) * 100)
        else:
            ocupacion_promedio = 0

        # Variación vs mes anterior
        mes_anterior_inicio = fecha_inicio - timedelta(days=30)
        inventario_mes_anterior = InventarioParqueo.objects.filter(
            parHoraEntrada__gte=mes_anterior_inicio,
            parHoraEntrada__lt=fecha_inicio
        ).count()
        inventario_actual = inventario_qs.count()

        # Variación porcentual vs el período anterior del mismo largo (positivo = crecimiento)
        if inventario_mes_anterior > 0:
            variacion_ocupacion = int(((inventario_actual - inventario_mes_anterior) / inventario_mes_anterior) * 100)
        else:
            variacion_ocupacion = 0

        # Ingresos mensuales (suma de pagos)
        ingresos_totales = Pago.objects.filter(
            pagFechaPago__gte=fecha_inicio,
            pagEstado=Pago.EstadoChoices.PAGADO
        ).aggregate(
            total=Sum('pagMonto')
        )['total'] or Decimal('0')

        ingresos_mes_anterior = Pago.objects.filter(
            pagFechaPago__gte=mes_anterior_inicio,
            pagFechaPago__lt=fecha_inicio,
            pagEstado=Pago.EstadoChoices.PAGADO
        ).aggregate(
            total=Sum('pagMonto')
        )['total'] or Decimal('0')

        if ingresos_mes_anterior > 0:
            variacion_ingresos = int(((ingresos_totales - ingresos_mes_anterior) / ingresos_mes_anterior) * 100)
        else:
            variacion_ingresos = 0

        # Usuarios activos (vehículos diferentes que han ingresado)
        usuarios_activos = inventario_qs.values('fkIdVehiculo').distinct().count()

        usuarios_mes_anterior = InventarioParqueo.objects.filter(
            parHoraEntrada__gte=mes_anterior_inicio,
            parHoraEntrada__lt=fecha_inicio
        ).values('fkIdVehiculo').distinct().count()

        if usuarios_mes_anterior > 0:
            variacion_usuarios = int(((usuarios_activos - usuarios_mes_anterior) / usuarios_mes_anterior) * 100)
        else:
            variacion_usuarios = 0

        # Rotación diaria: cuántas veces en promedio se ocupa cada espacio al día
        dias_en_periodo = (now - fecha_inicio).days or 1
        rotacion_diaria = round(inventario_actual / total_espacios / dias_en_periodo, 1) if total_espacios > 0 else 0

        # ══════════════════════════════════════════
        # 2. TENDENCIA DE OCUPACIÓN DIARIA (por hora HOY)
        # ══════════════════════════════════════════

        # Obtener fecha local de hoy
        hoy_local = timezone.localtime(now).date()

        # Traer todos los registros del día actual
        registros_hoy = InventarioParqueo.objects.filter(
            parHoraEntrada__range=(
                timezone.make_aware(datetime.combine(hoy_local, datetime.min.time())),
                timezone.make_aware(datetime.combine(hoy_local, datetime.max.time()))
            )
        )

        # Contar por hora en Python (más eficiente)
        counts_por_hora = {h: 0 for h in range(6, 23)}
        for r in registros_hoy:
            h = timezone.localtime(r.parHoraEntrada).hour
            if 6 <= h < 23:
                counts_por_hora[h] += 1

        # Generar labels y data
        ocupacion_labels = []
        ocupacion_data = []
        for h in range(6, 23):
            ocupacion_labels.append(f"{h}:00")
            ocupacion_data.append(counts_por_hora[h])

        # ══════════════════════════════════════════
        # 3. INGRESOS MENSUALES (últimos 6 meses)
        # ══════════════════════════════════════════

        ingresos_labels = []
        ingresos_data_values = []

        for i in range(5, -1, -1):  # i=5 → mes más antiguo, i=0 → mes actual
            mes_inicio = now - timedelta(days=30 * (i + 1))
            mes_fin = now - timedelta(days=30 * i)

            ingresos_mes = Pago.objects.filter(
                pagFechaPago__gte=mes_inicio,
                pagFechaPago__lt=mes_fin,
                pagEstado=Pago.EstadoChoices.PAGADO
            ).aggregate(
                total=Sum('pagMonto')
            )['total'] or Decimal('0')

            # Nombre del mes
            mes_nombre = mes_inicio.strftime('%b')
            ingresos_labels.append(mes_nombre)
            ingresos_data_values.append(float(ingresos_mes))

        # ══════════════════════════════════════════
        # 4. DISTRIBUCIÓN DE USO (Pie Chart)
        # ══════════════════════════════════════════

        # Residentes (vehículos registrados no visitantes)
        residentes_count = inventario_qs.filter(
            fkIdVehiculo__fkIdUsuario__isnull=False
        ).count()

        # Visitantes
        visitantes_count = inventario_qs.filter(
            fkIdVehiculo__fkIdUsuario__isnull=True
        ).count()

        # Reservas
        reservas_count = Reserva.objects.filter(
            resHoraInicio__gte=fecha_inicio
        ).count()

        total_uso = residentes_count + visitantes_count + reservas_count

        if total_uso > 0:
            residentes_pct = int((residentes_count / total_uso) * 100)
            visitantes_pct = int((visitantes_count / total_uso) * 100)
            reservas_pct = 100 - residentes_pct - visitantes_pct  # Se calcula como resto para que sumen exactamente 100%
        else:
            residentes_pct = visitantes_pct = reservas_pct = 0

        # ══════════════════════════════════════════
        # 5. RESUMEN POR PISO
        # ══════════════════════════════════════════

        pisos = Piso.objects.filter(pisEstado=True).prefetch_related('espacios')
        resumen_pisos = []

        for piso in pisos:
            total_espacios_piso = piso.espacios.filter(
                espEstado__in=['DISPONIBLE', 'OCUPADO']
            ).count()

            ocupados_piso = piso.espacios.filter(espEstado='OCUPADO').count()

            if total_espacios_piso > 0:
                ocupacion_piso = int((ocupados_piso / total_espacios_piso) * 100)
            else:
                ocupacion_piso = 0

            # Ingresos del piso
            ingresos_piso = Pago.objects.filter(
                fkIdParqueo__fkIdEspacio__fkIdPiso=piso,
                pagFechaPago__gte=fecha_inicio,
                pagEstado=Pago.EstadoChoices.PAGADO
            ).aggregate(
                total=Sum('pagMonto')
            )['total'] or Decimal('0')

            resumen_pisos.append({
                'nombre': piso.pisNombre,
                'ocupacion': ocupacion_piso,
                'ingresos': ingresos_piso,
            })

        # ══════════════════════════════════════════
        # CONTEXTO
        # ══════════════════════════════════════════

        context = {
            'active_page': 'reportes',
            # Filtros
            'periodo': periodo,
            'piso_id': piso_id,
            'pisos': pisos,

            # Métricas principales
            'ocupacion_promedio': ocupacion_promedio,
            'variacion_ocupacion': variacion_ocupacion,
            'ingresos_totales': ingresos_totales,
            'variacion_ingresos': variacion_ingresos,
            'usuarios_activos': usuarios_activos,
            'variacion_usuarios': variacion_usuarios,
            'rotacion_diaria': rotacion_diaria,

            # Datos para gráficas (convertidos a JSON)
            'ocupacion_labels': json.dumps(ocupacion_labels),
            'ocupacion_data': json.dumps(ocupacion_data),
            'ingresos_labels': json.dumps(ingresos_labels),
            'ingresos_data': json.dumps(ingresos_data_values),

            # Distribución de uso
            'residentes_pct': residentes_pct,
            'visitantes_pct': visitantes_pct,
            'reservas_pct': reservas_pct,

            # Resumen por piso
            'resumen_pisos': resumen_pisos,
        }

        return render(request, 'admin_panel/reportes.html', context)


class ExportarPDFReportesView(AdminRequiredMixin, View):
    def get(self, request):
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.lib.enums import TA_RIGHT, TA_CENTER, TA_LEFT
        import io

        PURPLE      = colors.HexColor('#7c3aed')
        PURPLE_DARK = colors.HexColor('#5b21b6')
        ROW_ALT     = colors.HexColor('#f5f3ff')
        GRAY_LIGHT  = colors.HexColor('#f9fafb')
        GRAY_TEXT   = colors.HexColor('#6b7280')
        BORDER      = colors.HexColor('#e5e7eb')
        BLACK       = colors.HexColor('#111827')
        GREEN       = colors.HexColor('#16a34a')

        periodo = request.GET.get('periodo', 'mes')
        now = timezone.now()
        if periodo == 'semana':
            fecha_inicio = now - timedelta(days=7)
            label = 'Última Semana'
        elif periodo == 'año':
            fecha_inicio = now - timedelta(days=365)
            label = 'Último Año'
        else:
            fecha_inicio = now - timedelta(days=30)
            label = 'Último Mes'

        pagos = Pago.objects.filter(
            pagFechaPago__gte=fecha_inicio,
            pagEstado='PAGADO'
        ).select_related(
            'fkIdParqueo__fkIdVehiculo',
            'fkIdParqueo__fkIdEspacio__fkIdPiso'
        ).order_by('-pagFechaPago')

        total = pagos.aggregate(t=Sum('pagMonto'))['t'] or 0

        buffer = io.BytesIO()
        W, H = A4
        doc = SimpleDocTemplate(
            buffer, pagesize=A4,
            topMargin=0, bottomMargin=20*mm,
            leftMargin=18*mm, rightMargin=18*mm
        )

        s_title   = ParagraphStyle('title',   fontName='Helvetica-Bold', fontSize=20, textColor=colors.white,   leading=24, alignment=TA_LEFT)
        s_sub     = ParagraphStyle('sub',     fontName='Helvetica',      fontSize=9,  textColor=colors.HexColor('#ddd6fe'), leading=12, alignment=TA_LEFT)
        s_label   = ParagraphStyle('label',   fontName='Helvetica',      fontSize=8,  textColor=GRAY_TEXT, leading=10, spaceAfter=2)
        s_value   = ParagraphStyle('value',   fontName='Helvetica-Bold', fontSize=14, textColor=BLACK,    leading=16)
        s_value_g = ParagraphStyle('value_g', fontName='Helvetica-Bold', fontSize=16, textColor=GREEN,    leading=18)
        s_footer  = ParagraphStyle('footer',  fontName='Helvetica',      fontSize=7,  textColor=GRAY_TEXT, alignment=TA_CENTER)

        PAGE_W = W - 36*mm  # usable width

        elements = []

        # ── HEADER BAND (purple box as table trick) ──────────────────────
        now_local = timezone.localtime(now)
        header_data = [[
            Paragraph(f'MultiParking', s_title),
            Paragraph(f'Reporte de Ingresos — {label}<br/><font size="8" color="#ddd6fe">Generado: {now_local.strftime("%d/%m/%Y %H:%M")}</font>', s_sub),
        ]]
        header_table = Table(header_data, colWidths=[PAGE_W * 0.4, PAGE_W * 0.6])
        header_table.setStyle(TableStyle([
            ('BACKGROUND',   (0, 0), (-1, -1), PURPLE_DARK),
            ('TOPPADDING',   (0, 0), (-1, -1), 14),
            ('BOTTOMPADDING',(0, 0), (-1, -1), 14),
            ('LEFTPADDING',  (0, 0), (-1, -1), 14),
            ('RIGHTPADDING', (0, 0), (-1, -1), 14),
            ('VALIGN',       (0, 0), (-1, -1), 'MIDDLE'),
            ('ROUNDEDCORNERS', [6]),
        ]))
        elements.append(header_table)
        elements.append(Spacer(1, 14))

        # ── SUMMARY CARDS ────────────────────────────────────────────────
        summary_data = [[
            Paragraph('Total Recaudado', s_label),
            Paragraph('Pagos Incluidos', s_label),
            Paragraph('Período', s_label),
        ], [
            Paragraph(f'${float(total):,.0f} COP', s_value_g),
            Paragraph(str(pagos.count()), s_value),
            Paragraph(label, s_value),
        ]]
        col_w = PAGE_W / 3
        summary_table = Table(summary_data, colWidths=[col_w, col_w, col_w])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND',    (0, 0), (-1, -1), GRAY_LIGHT),
            ('TOPPADDING',    (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
            ('LEFTPADDING',   (0, 0), (-1, -1), 12),
            ('RIGHTPADDING',  (0, 0), (-1, -1), 12),
            ('BOX',           (0, 0), (-1, -1), 0.5, BORDER),
            ('LINEAFTER',     (0, 0), (1, -1),  0.5, BORDER),
            ('ROUNDEDCORNERS', [4]),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 16))

        # ── TABLE ────────────────────────────────────────────────────────
        col_widths = [28*mm, 22*mm, 28*mm, 22*mm, 24*mm, 30*mm]
        data = [[
            Paragraph('<b>Fecha</b>',   ParagraphStyle('h', fontName='Helvetica-Bold', fontSize=8, textColor=colors.white)),
            Paragraph('<b>Hora</b>',    ParagraphStyle('h', fontName='Helvetica-Bold', fontSize=8, textColor=colors.white)),
            Paragraph('<b>Placa</b>',   ParagraphStyle('h', fontName='Helvetica-Bold', fontSize=8, textColor=colors.white)),
            Paragraph('<b>Espacio</b>', ParagraphStyle('h', fontName='Helvetica-Bold', fontSize=8, textColor=colors.white)),
            Paragraph('<b>Método</b>',  ParagraphStyle('h', fontName='Helvetica-Bold', fontSize=8, textColor=colors.white)),
            Paragraph('<b>Monto COP</b>', ParagraphStyle('h', fontName='Helvetica-Bold', fontSize=8, textColor=colors.white, alignment=TA_RIGHT)),
        ]]

        s_cell = ParagraphStyle('cell', fontName='Helvetica', fontSize=8, textColor=BLACK, leading=10)
        s_mono = ParagraphStyle('mono', fontName='Helvetica-Bold', fontSize=8, textColor=BLACK, leading=10)
        s_amt  = ParagraphStyle('amt',  fontName='Helvetica-Bold', fontSize=8, textColor=GREEN, leading=10, alignment=TA_RIGHT)

        for i, p in enumerate(pagos[:500]):
            fecha_local = timezone.localtime(p.pagFechaPago)
            piso = ''
            try:
                piso = p.fkIdParqueo.fkIdEspacio.fkIdPiso.pisNombre
            except Exception:
                pass
            data.append([
                Paragraph(fecha_local.strftime('%d/%m/%Y'), s_cell),
                Paragraph(fecha_local.strftime('%H:%M'),    s_cell),
                Paragraph(p.fkIdParqueo.fkIdVehiculo.vehPlaca, s_mono),
                Paragraph(f"{p.fkIdParqueo.fkIdEspacio.espNumero}<br/><font size='7' color='#9ca3af'>{piso}</font>", s_cell),
                Paragraph(p.pagMetodo, s_cell),
                Paragraph(f'${float(p.pagMonto):,.0f}', s_amt),
            ])

        # totals row
        data.append([
            Paragraph('<b>TOTAL</b>', ParagraphStyle('tot', fontName='Helvetica-Bold', fontSize=8, textColor=colors.white)),
            '', '', '', '',
            Paragraph(f'<b>${float(total):,.0f}</b>', ParagraphStyle('tamt', fontName='Helvetica-Bold', fontSize=8, textColor=colors.white, alignment=TA_RIGHT)),
        ])

        n = len(data)
        row_styles = []
        for r in range(1, n - 1):
            bg = ROW_ALT if r % 2 == 0 else colors.white
            row_styles.append(('BACKGROUND', (0, r), (-1, r), bg))

        t = Table(data, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            # header
            ('BACKGROUND',    (0, 0), (-1, 0),  PURPLE),
            ('TOPPADDING',    (0, 0), (-1, 0),  8),
            ('BOTTOMPADDING', (0, 0), (-1, 0),  8),
            # cells
            ('TOPPADDING',    (0, 1), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 5),
            ('LEFTPADDING',   (0, 0), (-1, -1), 6),
            ('RIGHTPADDING',  (0, 0), (-1, -1), 6),
            # totals row
            ('BACKGROUND',    (0, n-1), (-1, n-1), PURPLE_DARK),
            ('SPAN',          (0, n-1), (4, n-1)),
            # borders
            ('LINEBELOW',     (0, 0), (-1, 0),  0.5, PURPLE_DARK),
            ('LINEBELOW',     (0, 1), (-1, -2), 0.3, BORDER),
            ('BOX',           (0, 0), (-1, -1), 0.5, BORDER),
            ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
        ] + row_styles))
        elements.append(t)
        elements.append(Spacer(1, 10))

        # ── FOOTER ───────────────────────────────────────────────────────
        elements.append(HRFlowable(width='100%', thickness=0.5, color=BORDER))
        elements.append(Spacer(1, 4))
        elements.append(Paragraph(
            f'MultiParking · Reporte generado el {now_local.strftime("%d/%m/%Y a las %H:%M")} · Solo pagos con estado PAGADO',
            s_footer
        ))

        def add_page_num(canvas, doc):
            canvas.saveState()
            canvas.setFont('Helvetica', 7)
            canvas.setFillColor(GRAY_TEXT)
            canvas.drawRightString(W - 18*mm, 12*mm, f'Página {doc.page}')
            canvas.restoreState()

        doc.build(elements, onFirstPage=add_page_num, onLaterPages=add_page_num)
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="reporte_ingresos_{periodo}.pdf"'
        return response


class ExportarExcelReportesView(AdminRequiredMixin, View):
    def get(self, request):
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        periodo = request.GET.get('periodo', 'mes')
        now = timezone.now()
        if periodo == 'semana':
            fecha_inicio = now - timedelta(days=7)
            label = 'semana'
        elif periodo == 'año':
            fecha_inicio = now - timedelta(days=365)
            label = 'anio'
        else:
            fecha_inicio = now - timedelta(days=30)
            label = 'mes'

        pagos = Pago.objects.filter(
            pagFechaPago__gte=fecha_inicio,
            pagEstado='PAGADO'
        ).select_related('fkIdParqueo__fkIdVehiculo', 'fkIdParqueo__fkIdEspacio__fkIdPiso')

        wb = Workbook()
        ws = wb.active
        ws.title = 'Ingresos'

        header_fill = PatternFill(start_color='6D28D9', end_color='6D28D9', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)

        headers = ['Fecha', 'Hora', 'Placa', 'Espacio', 'Piso', 'Método', 'Monto COP']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        for row, p in enumerate(pagos, 2):
            ws.cell(row=row, column=1, value=p.pagFechaPago.strftime('%d/%m/%Y'))
            ws.cell(row=row, column=2, value=p.pagFechaPago.strftime('%H:%M'))
            ws.cell(row=row, column=3, value=p.fkIdParqueo.fkIdVehiculo.vehPlaca)
            ws.cell(row=row, column=4, value=p.fkIdParqueo.fkIdEspacio.espNumero)
            ws.cell(row=row, column=5, value=p.fkIdParqueo.fkIdEspacio.fkIdPiso.pisNombre)
            ws.cell(row=row, column=6, value=p.pagMetodo)
            ws.cell(row=row, column=7, value=float(p.pagMonto))

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 15

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(
            buffer,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="reporte_ingresos_{label}.xlsx"'
        return response
